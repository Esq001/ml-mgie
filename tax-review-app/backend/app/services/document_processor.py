import base64
import io
import os
import uuid
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl
import pandas as pd
from PIL import Image

from app.config import settings
from app.models.schemas import DocumentType, UploadedDocument


class DocumentProcessor:
    """Handles extraction of text and data from tax documents."""

    SUPPORTED_EXTENSIONS = {
        ".pdf": "pdf",
        ".xlsx": "excel",
        ".xls": "excel",
        ".csv": "csv",
        ".png": "image",
        ".jpg": "image",
        ".jpeg": "image",
        ".tiff": "image",
        ".tif": "image",
    }

    def __init__(self):
        os.makedirs(settings.upload_dir, exist_ok=True)

    async def process_upload(
        self, filename: str, content: bytes
    ) -> UploadedDocument:
        ext = Path(filename).suffix.lower()
        if ext not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(
                f"Unsupported file type: {ext}. "
                f"Supported: {', '.join(self.SUPPORTED_EXTENSIONS.keys())}"
            )

        doc_id = str(uuid.uuid4())
        file_type = self.SUPPORTED_EXTENSIONS[ext]

        # Save file
        save_path = Path(settings.upload_dir) / f"{doc_id}{ext}"
        save_path.write_bytes(content)

        # Extract text based on type
        extracted_text = ""
        page_count = None

        if file_type == "pdf":
            extracted_text, page_count = self._extract_pdf(content)
        elif file_type == "excel":
            extracted_text = self._extract_excel(content, ext)
        elif file_type == "csv":
            extracted_text = self._extract_csv(content)
        elif file_type == "image":
            extracted_text = self._prepare_image_description(content, ext)

        doc_type = self._classify_document(filename, extracted_text)

        return UploadedDocument(
            id=doc_id,
            filename=filename,
            file_type=file_type,
            document_type=doc_type,
            size_bytes=len(content),
            page_count=page_count,
            extracted_text=extracted_text,
        )

    def _extract_pdf(self, content: bytes) -> tuple[str, int]:
        doc = fitz.open(stream=content, filetype="pdf")
        pages = []
        for page_num, page in enumerate(doc, 1):
            text = page.get_text()
            if text.strip():
                pages.append(f"--- Page {page_num} ---\n{text}")
            else:
                # Page might be scanned — extract as image for Claude vision
                pix = page.get_pixmap(dpi=200)
                img_bytes = pix.tobytes("png")
                b64 = base64.b64encode(img_bytes).decode()
                pages.append(
                    f"--- Page {page_num} (scanned) ---\n"
                    f"[IMAGE:data:image/png;base64,{b64}]"
                )
        page_count = len(doc)
        doc.close()
        return "\n\n".join(pages), page_count

    def _extract_excel(self, content: bytes, ext: str) -> str:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheets_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_str = "\t".join(
                    str(cell) if cell is not None else "" for cell in row
                )
                if row_str.strip():
                    rows.append(row_str)
            if rows:
                sheets_text.append(
                    f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows)
                )
        wb.close()
        return "\n\n".join(sheets_text)

    def _extract_csv(self, content: bytes) -> str:
        df = pd.read_csv(io.BytesIO(content))
        return f"=== CSV Data ===\n{df.to_string()}"

    def _prepare_image_description(self, content: bytes, ext: str) -> str:
        img = Image.open(io.BytesIO(content))
        # Convert to PNG for consistent handling
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        b64 = base64.b64encode(buf.getvalue()).decode()
        return f"[IMAGE:data:image/png;base64,{b64}]"

    def _classify_document(self, filename: str, text: str) -> DocumentType:
        name_lower = filename.lower()
        text_lower = text[:2000].lower()

        if any(
            kw in name_lower
            for kw in ["1040", "1120", "1065", "990", "return"]
        ):
            return DocumentType.TAX_RETURN
        if any(kw in name_lower for kw in ["workpaper", "work paper", "wp"]):
            return DocumentType.WORK_PAPER
        if any(
            kw in name_lower
            for kw in ["schedule", "sch", "k-1", "k1"]
        ):
            return DocumentType.SCHEDULE

        if any(
            kw in text_lower
            for kw in [
                "form 1040",
                "form 1120",
                "form 1065",
                "u.s. income tax return",
                "tax return",
            ]
        ):
            return DocumentType.TAX_RETURN
        if any(
            kw in text_lower
            for kw in ["work paper", "workpaper", "adjusting entry"]
        ):
            return DocumentType.WORK_PAPER

        return DocumentType.OTHER

    def get_image_blocks(self, extracted_text: str) -> list[dict]:
        """Extract base64 image blocks for Claude vision API."""
        blocks = []
        for segment in extracted_text.split("[IMAGE:"):
            if segment.startswith("data:image"):
                end = segment.index("]")
                data_uri = segment[:end]
                media_type, b64_data = data_uri.split(";base64,")
                media_type = media_type.replace("data:", "")
                blocks.append(
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": b64_data,
                        },
                    }
                )
        return blocks

    def get_text_content(self, extracted_text: str) -> str:
        """Get text-only content, stripping image blocks."""
        import re
        return re.sub(r"\[IMAGE:data:image/[^\]]+\]", "[scanned page]", extracted_text)
