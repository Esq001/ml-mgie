"""
Excel (.xlsx) output writer for K-1 extracted data.

Generates a formatted workbook with structured sheets:
1. K-1 Data - Clean table with only populated columns, grouped headers
2. Per-Recipient Detail - One section per recipient with vertical layout
3. Summary - Aggregated totals and form counts
4. Extraction Log - Per-file processing details
"""

import logging
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from .models import K1Data
from .constants import FORM_DEFINITIONS, ALL_BOX_IDS

logger = logging.getLogger(__name__)

# =============================================================================
# Styling
# =============================================================================

HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

GROUP_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
GROUP_FILL_INFO = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
GROUP_FILL_INCOME = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
GROUP_FILL_DEDUCT = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
GROUP_FILL_OTHER = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

SUB_HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="333333")
SUB_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUB_HEADER_ALIGN = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

ROW_EVEN_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
CURRENCY_FORMAT = '#,##0.00;(#,##0.00);"-"'
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="2E75B6")
LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="333333")
LABEL_FILL = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")

HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LOW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

CONFIDENCE_FILLS = {"high": HIGH_FILL, "medium": MED_FILL, "low": LOW_FILL}


def write_excel(data: List[K1Data], output_path: str) -> None:
    """Write extracted K-1 data to a formatted Excel workbook."""
    wb = Workbook()

    _write_data_sheet(wb, data)
    _write_detail_sheet(wb, data)
    _write_summary_sheet(wb, data)
    _write_log_sheet(wb, data)

    wb.save(output_path)
    logger.info("Excel file saved to: %s", output_path)


# =============================================================================
# Sheet 1: K-1 Data (clean table with only populated box columns)
# =============================================================================

def _write_data_sheet(wb: Workbook, data: List[K1Data]) -> None:
    """Write the main data table — only includes box columns that have data."""
    ws = wb.active
    ws.title = "K-1 Data"

    # Determine which box IDs actually have data across all forms
    used_box_ids = []
    for box_id in ALL_BOX_IDS:
        if any(box_id in k1.boxes for k1 in data):
            used_box_ids.append(box_id)

    # ---- Row 1: Group headers ----
    info_cols = ["Source File", "Form", "Entity Name", "EIN",
                 "Tax Year", "Recipient", "Recipient ID", "Type", "Confidence"]
    info_count = len(info_cols)

    # Categorize boxes into groups for coloring
    income_boxes = [b for b in used_box_ids if _box_number(b) <= 11]
    deduction_boxes = [b for b in used_box_ids if 12 <= _box_number(b) <= 13]
    other_boxes = [b for b in used_box_ids if _box_number(b) >= 14]

    # Write group header row (Row 1)
    # Info group
    for col in range(1, info_count + 1):
        cell = ws.cell(row=1, column=col)
        if col == 1:
            cell.value = "Form Information"
        cell.font = GROUP_FONT
        cell.fill = GROUP_FILL_INFO
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    if info_count > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=info_count)

    col_offset = info_count + 1
    for group_boxes, group_label, group_fill in [
        (income_boxes, "Income / Gains", GROUP_FILL_INCOME),
        (deduction_boxes, "Deductions / Credits", GROUP_FILL_DEDUCT),
        (other_boxes, "Other Items", GROUP_FILL_OTHER),
    ]:
        if not group_boxes:
            continue
        start_col = col_offset
        for i, _ in enumerate(group_boxes):
            cell = ws.cell(row=1, column=col_offset + i)
            if i == 0:
                cell.value = group_label
            cell.font = GROUP_FONT
            cell.fill = group_fill
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
        if len(group_boxes) > 1:
            ws.merge_cells(start_row=1, start_column=start_col,
                           end_row=1, end_column=start_col + len(group_boxes) - 1)
        col_offset += len(group_boxes)

    # ---- Row 2: Column sub-headers ----
    for col_idx, name in enumerate(info_cols, start=1):
        cell = ws.cell(row=2, column=col_idx, value=name)
        cell.font = SUB_HEADER_FONT
        cell.fill = SUB_HEADER_FILL
        cell.alignment = SUB_HEADER_ALIGN
        cell.border = THIN_BORDER

    col_offset = info_count + 1
    for box_id in used_box_ids:
        label = _short_box_label(box_id)
        cell = ws.cell(row=2, column=col_offset, value=label)
        cell.font = SUB_HEADER_FONT
        cell.fill = SUB_HEADER_FILL
        cell.alignment = SUB_HEADER_ALIGN
        cell.border = THIN_BORDER
        col_offset += 1

    # Freeze first 2 rows + first column
    ws.freeze_panes = "B3"

    # Set row 2 height for wrapped headers
    ws.row_dimensions[2].height = 45

    # ---- Data rows (Row 3+) ----
    total_cols = info_count + len(used_box_ids)
    for row_idx, k1 in enumerate(data, start=3):
        is_even = (row_idx - 3) % 2 == 1

        info_values = [
            k1.source_file,
            k1.form_type,
            k1.entity_name,
            k1.entity_ein,
            k1.tax_year,
            k1.recipient_name,
            k1.recipient_id,
            k1.recipient_type,
            k1.confidence,
        ]

        for col_idx, value in enumerate(info_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=10)
            if is_even:
                cell.fill = ROW_EVEN_FILL
            # Color confidence
            if col_idx == len(info_values) and value in CONFIDENCE_FILLS:
                cell.fill = CONFIDENCE_FILLS[value]

        # Box columns
        col_offset = info_count + 1
        for box_id in used_box_ids:
            cell = ws.cell(row=row_idx, column=col_offset)
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=10)
            if is_even:
                cell.fill = ROW_EVEN_FILL

            raw = k1.boxes.get(box_id, "")
            if raw:
                try:
                    cell.value = float(raw.replace(",", ""))
                    cell.number_format = CURRENCY_FORMAT
                    cell.alignment = Alignment(horizontal="right")
                except (ValueError, TypeError):
                    cell.value = raw
            col_offset += 1

    # Auto-filter & column widths
    if data:
        ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{len(data) + 2}"
    _auto_fit_columns(ws, min_width=8, max_width=22)

    # Widen key columns
    ws.column_dimensions["A"].width = 28  # Source File
    ws.column_dimensions["C"].width = 25  # Entity Name
    ws.column_dimensions["F"].width = 22  # Recipient


# =============================================================================
# Sheet 2: Per-Recipient Detail (vertical card layout)
# =============================================================================

def _write_detail_sheet(wb: Workbook, data: List[K1Data]) -> None:
    """Write a detail sheet with one vertical section per K-1 form."""
    ws = wb.create_sheet("Per-Recipient Detail")

    row = 1
    ws.cell(row=row, column=1, value="K-1 Per-Recipient Detail").font = TITLE_FONT
    ws.merge_cells("A1:D1")
    row = 3

    for idx, k1 in enumerate(data):
        # Section header
        header_text = f"K-1 #{idx + 1}: {k1.recipient_name or 'Unknown'}"
        cell = ws.cell(row=row, column=1, value=header_text)
        cell.font = SECTION_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        # Divider line
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = Border(
                bottom=Side(style="medium", color="2E75B6"))
        row += 1

        # Form info section
        info_fields = [
            ("Source File", k1.source_file),
            ("Form Type", _form_type_display(k1.form_type)),
            ("Confidence", k1.confidence),
            ("Entity Name", k1.entity_name),
            ("Entity EIN", k1.entity_ein),
            ("Tax Year", k1.tax_year),
            ("Recipient Name", k1.recipient_name),
            ("Recipient ID", k1.recipient_id),
            ("Recipient Type", k1.recipient_type),
        ]

        for label, value in info_fields:
            if not value:
                continue
            lcell = ws.cell(row=row, column=1, value=label)
            lcell.font = LABEL_FONT
            lcell.fill = LABEL_FILL
            lcell.border = THIN_BORDER
            lcell.alignment = Alignment(horizontal="right")

            vcell = ws.cell(row=row, column=2, value=value)
            vcell.border = THIN_BORDER
            vcell.font = Font(name="Calibri", size=10)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

            if label == "Confidence" and value in CONFIDENCE_FILLS:
                vcell.fill = CONFIDENCE_FILLS[value]
            row += 1

        row += 1

        # Box values table
        if k1.boxes:
            # Header
            for col_idx, hdr in enumerate(["Box", "Description", "Amount"], start=1):
                cell = ws.cell(row=row, column=col_idx, value=hdr)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
                cell.border = THIN_BORDER
            row += 1

            box_defs = FORM_DEFINITIONS.get(k1.form_type, FORM_DEFINITIONS["1065"])["boxes"]
            sorted_boxes = sorted(
                k1.boxes.items(),
                key=lambda x: (_box_number(x[0]), x[0])
            )

            for i, (box_id, raw_value) in enumerate(sorted_boxes):
                is_even = i % 2 == 1
                desc = box_defs.get(box_id, "")

                ws.cell(row=row, column=1, value=f"Box {box_id}").border = THIN_BORDER
                ws.cell(row=row, column=2, value=desc).border = THIN_BORDER
                amount_cell = ws.cell(row=row, column=3)
                amount_cell.border = THIN_BORDER

                if is_even:
                    for c in range(1, 4):
                        ws.cell(row=row, column=c).fill = ROW_EVEN_FILL

                try:
                    amount_cell.value = float(raw_value.replace(",", ""))
                    amount_cell.number_format = CURRENCY_FORMAT
                    amount_cell.alignment = Alignment(horizontal="right")
                except (ValueError, TypeError):
                    amount_cell.value = raw_value

                row += 1

        row += 2  # Gap between K-1 sections

    _auto_fit_columns(ws, min_width=12, max_width=40)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18


# =============================================================================
# Sheet 3: Summary
# =============================================================================

def _write_summary_sheet(wb: Workbook, data: List[K1Data]) -> None:
    """Write a summary sheet with totals and counts."""
    ws = wb.create_sheet("Summary")

    ws.cell(row=1, column=1, value="K-1 Extraction Summary").font = TITLE_FONT
    ws.merge_cells("A1:D1")

    row = 3

    # --- Forms Processed ---
    ws.cell(row=row, column=1, value="Forms Processed").font = SECTION_FONT
    row += 1

    type_counts = {}
    for k1 in data:
        type_counts[k1.form_type] = type_counts.get(k1.form_type, 0) + 1

    for hdr, col in [("Form Type", 1), ("Count", 2)]:
        cell = ws.cell(row=row, column=col, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for form_type, count in sorted(type_counts.items()):
        ws.cell(row=row, column=1, value=_form_type_display(form_type)).border = THIN_BORDER
        ws.cell(row=row, column=2, value=count).border = THIN_BORDER
        row += 1

    cell = ws.cell(row=row, column=1, value="Total")
    cell.font = Font(bold=True)
    cell.border = THIN_BORDER
    cell = ws.cell(row=row, column=2, value=len(data))
    cell.font = Font(bold=True)
    cell.border = THIN_BORDER

    row += 2

    # --- Confidence ---
    ws.cell(row=row, column=1, value="Extraction Confidence").font = SECTION_FONT
    row += 1

    for hdr, col in [("Level", 1), ("Count", 2)]:
        cell = ws.cell(row=row, column=col, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    conf_counts = {"high": 0, "medium": 0, "low": 0}
    for k1 in data:
        conf_counts[k1.confidence] = conf_counts.get(k1.confidence, 0) + 1

    for level in ["high", "medium", "low"]:
        cell = ws.cell(row=row, column=1, value=level.capitalize())
        cell.fill = CONFIDENCE_FILLS[level]
        cell.border = THIN_BORDER
        ws.cell(row=row, column=2, value=conf_counts[level]).border = THIN_BORDER
        row += 1

    row += 2

    # --- Aggregate Totals ---
    ws.cell(row=row, column=1, value="Aggregate Box Totals").font = SECTION_FONT
    row += 1

    for hdr, col in [("Box", 1), ("Description", 2), ("Total", 3), ("# Forms", 4)]:
        cell = ws.cell(row=row, column=col, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for box_id in ALL_BOX_IDS:
        total = 0.0
        count = 0
        for k1 in data:
            val = k1.box_value_as_float(box_id)
            if val is not None:
                total += val
                count += 1

        if count == 0:
            continue

        label = ""
        for form_def in FORM_DEFINITIONS.values():
            if box_id in form_def["boxes"]:
                label = form_def["boxes"][box_id]
                break

        is_even = (row % 2) == 0
        ws.cell(row=row, column=1, value=f"Box {box_id}").border = THIN_BORDER
        ws.cell(row=row, column=2, value=label).border = THIN_BORDER
        total_cell = ws.cell(row=row, column=3, value=total)
        total_cell.number_format = CURRENCY_FORMAT
        total_cell.border = THIN_BORDER
        ws.cell(row=row, column=4, value=count).border = THIN_BORDER

        if is_even:
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = ROW_EVEN_FILL
        row += 1

    _auto_fit_columns(ws, min_width=10, max_width=40)
    ws.column_dimensions["B"].width = 38


# =============================================================================
# Sheet 4: Extraction Log
# =============================================================================

def _write_log_sheet(wb: Workbook, data: List[K1Data]) -> None:
    """Write the extraction log with per-file details."""
    ws = wb.create_sheet("Extraction Log")

    headers = ["Source File", "Form Type", "Method", "Confidence",
               "Boxes Found", "Warnings"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"

    for row_idx, k1 in enumerate(data, start=2):
        is_even = (row_idx - 2) % 2 == 1
        values = [
            k1.source_file,
            _form_type_display(k1.form_type),
            k1.extraction_method.upper() if k1.extraction_method else "",
            k1.confidence,
            len(k1.boxes),
            "; ".join(k1.warnings) if k1.warnings else "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if is_even:
                cell.fill = ROW_EVEN_FILL
            if col_idx == 4 and value in CONFIDENCE_FILLS:
                cell.fill = CONFIDENCE_FILLS[value]

    _auto_fit_columns(ws, min_width=10, max_width=50)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["F"].width = 45


# =============================================================================
# Helpers
# =============================================================================

def _auto_fit_columns(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Auto-fit column widths based on cell content."""
    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width


def _short_box_label(box_id: str) -> str:
    """Get a short column header for a box: 'Box 1\nOrd. business income'."""
    label = ""
    for form_def in FORM_DEFINITIONS.values():
        if box_id in form_def["boxes"]:
            label = form_def["boxes"][box_id]
            break
    # Shorten label
    short = label[:28] + "..." if len(label) > 30 else label
    return f"Box {box_id}\n{short}"


def _box_number(box_id: str) -> int:
    """Extract the numeric part of a box ID for sorting (e.g., '9a' -> 9)."""
    digits = "".join(c for c in box_id if c.isdigit())
    return int(digits) if digits else 0


def _form_type_display(form_type: str) -> str:
    """Return a display-friendly form type string."""
    return {
        "1065": "1065 (Partnership)",
        "1120S": "1120-S (S-Corp)",
        "1041": "1041 (Estate/Trust)",
    }.get(form_type, form_type)
